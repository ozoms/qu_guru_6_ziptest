import pathlib
import zipfile
import os
import shutil
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook

current_dir = os.path.abspath(os.path.dirname(__file__))
resources_dir = os.path.join(current_dir, '../resources')
directory = pathlib.Path(resources_dir)

with zipfile.ZipFile('testzip.zip', mode="w") as archive:
    for file_path in directory.iterdir():
        archive.write(file_path, arcname=file_path.name)

shutil.move('testzip.zip', directory)

resources_zip_dir = os.path.join(resources_dir, 'testzip.zip')

zipfile.ZipFile(resources_zip_dir).extractall()

text = PdfReader('tests_doc.pdf').pages[0].extract_text()
assert '1 a' in text

with open('tests_data.csv') as csvfile:
    csvfile = csv.reader(csvfile)
    assert ['1', 'a'] in csvfile

workbook = load_workbook('tests_table.xlsx')
sheet = workbook.active
assert 'a' == sheet.cell(row=1, column=2).value